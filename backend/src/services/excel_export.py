"""
Servicio para exportación de facturas a Excel.
Genera archivos Excel con las facturas filtradas.
"""

import os
from datetime import datetime
from typing import List
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.models import Invoice


def export_invoices_to_excel(invoices: List[Invoice]) -> str:
    """
    Exportar facturas a archivo Excel con formato profesional.
    
    Args:
        invoices: Lista de facturas a exportar
        
    Returns:
        str: Ruta del archivo Excel generado
    """
    # Crear workbook y worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Facturas Boosting"
    
    # Definir estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # Encabezados
    headers = [
        "ID", "Colaborador", "Fecha", "Proveedor", "Monto", 
        "Método de Pago", "Categoría", "Estado", "Descripción", "Fecha de Registro"
    ]
    
    # Escribir encabezados
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    # Escribir datos
    for row, invoice in enumerate(invoices, 2):
        ws.cell(row=row, column=1, value=invoice.id).border = border
        ws.cell(row=row, column=2, value=invoice.user.name).border = border
        ws.cell(row=row, column=3, value=invoice.date.strftime("%Y-%m-%d")).border = border
        ws.cell(row=row, column=4, value=invoice.provider).border = border
        ws.cell(row=row, column=5, value=f"${invoice.amount:,.2f}").border = border
        ws.cell(row=row, column=6, value=invoice.payment_method.value).border = border
        ws.cell(row=row, column=7, value=invoice.category.value).border = border
        ws.cell(row=row, column=8, value=invoice.status.value).border = border
        ws.cell(row=row, column=9, value=invoice.description or "").border = border
        ws.cell(row=row, column=10, value=invoice.created_at.strftime("%Y-%m-%d %H:%M")).border = border
    
    # Ajustar ancho de columnas
    column_widths = [8, 20, 12, 25, 12, 15, 15, 12, 30, 18]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Agregar información del reporte
    ws.insert_rows(1)
    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    title_cell = ws.cell(row=1, column=1, value="REPORTE DE FACTURAS - BOOSTING")
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = center_alignment
    title_cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    
    # Agregar fecha de generación
    ws.insert_rows(2)
    ws.merge_cells(f"A2:{get_column_letter(len(headers))}2")
    date_cell = ws.cell(row=2, column=1, value=f"Generado el: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    date_cell.alignment = center_alignment
    date_cell.font = Font(italic=True)
    
    # Agregar totales
    total_row = len(invoices) + 4
    ws.cell(row=total_row, column=4, value="TOTAL:").font = Font(bold=True)
    ws.cell(row=total_row, column=5, value=f"${sum(invoice.amount for invoice in invoices):,.2f}").font = Font(bold=True)
    
    # Generar nombre de archivo único
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"facturas_boosting_{timestamp}.xlsx"
    filepath = os.path.join("exports", filename)
    
    # Crear directorio de exports si no existe
    os.makedirs("exports", exist_ok=True)
    
    # Guardar archivo
    wb.save(filepath)
    
    return filepath


def export_invoices_summary_to_excel(invoices: List[Invoice]) -> str:
    """
    Exportar resumen de facturas por usuario y categoría.
    
    Args:
        invoices: Lista de facturas a exportar
        
    Returns:
        str: Ruta del archivo Excel generado
    """
    wb = openpyxl.Workbook()
    
    # Hoja 1: Resumen por usuario
    ws_users = wb.active
    ws_users.title = "Resumen por Usuario"
    
    # Agrupar por usuario
    user_totals = {}
    for invoice in invoices:
        user_name = invoice.user.name
        if user_name not in user_totals:
            user_totals[user_name] = 0
        user_totals[user_name] += invoice.amount
    
    # Escribir resumen por usuario
    headers = ["Usuario", "Total Facturado", "Número de Facturas"]
    for col, header in enumerate(headers, 1):
        ws_users.cell(row=1, column=col, value=header)
    
    row = 2
    for user_name, total in user_totals.items():
        count = len([inv for inv in invoices if inv.user.name == user_name])
        ws_users.cell(row=row, column=1, value=user_name)
        ws_users.cell(row=row, column=2, value=f"${total:,.2f}")
        ws_users.cell(row=row, column=3, value=count)
        row += 1
    
    # Hoja 2: Resumen por categoría
    ws_categories = wb.create_sheet("Resumen por Categoría")
    
    # Agrupar por categoría
    category_totals = {}
    for invoice in invoices:
        category = invoice.category.value
        if category not in category_totals:
            category_totals[category] = 0
        category_totals[category] += invoice.amount
    
    # Escribir resumen por categoría
    headers = ["Categoría", "Total Facturado", "Número de Facturas"]
    for col, header in enumerate(headers, 1):
        ws_categories.cell(row=1, column=col, value=header)
    
    row = 2
    for category, total in category_totals.items():
        count = len([inv for inv in invoices if inv.category.value == category])
        ws_categories.cell(row=row, column=1, value=category)
        ws_categories.cell(row=row, column=2, value=f"${total:,.2f}")
        ws_categories.cell(row=row, column=3, value=count)
        row += 1
    
    # Generar nombre de archivo único
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"resumen_facturas_boosting_{timestamp}.xlsx"
    filepath = os.path.join("exports", filename)
    
    # Crear directorio de exports si no existe
    os.makedirs("exports", exist_ok=True)
    
    # Guardar archivo
    wb.save(filepath)
    
    return filepath
